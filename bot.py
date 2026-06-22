"""
Securities Helper Bot - Telegram bot for managing security prices and client reports.
Integrates with Excel workbooks for data management and web scraping for price updates.
"""

import os
import logging
from typing import List, Optional
from datetime import date, datetime
from concurrent.futures import ThreadPoolExecutor
import re

import telebot
from telebot import types
from flask import Flask, request
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv

# ============================================================================
# CONFIGURATION
# ============================================================================
load_dotenv()

# Environment variables
TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
ADMIN_ID = int(os.getenv('ADMIN_CHAT_ID', 0))
WHITELIST = list(map(int, os.getenv('WHITELIST_IDS', '').split(','))) if os.getenv('WHITELIST_IDS') else []
DATA_DIR = os.getenv('DATA_DIR', 'data')
DB_FILE = os.path.join(DATA_DIR, 'auto_base.xlsx')
MAIN_SHEET = 'Список бумаг'
MAX_WORKERS = int(os.getenv('MAX_WORKERS', 8))

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize bot and server
bot = telebot.TeleBot(token=TOKEN)
server = Flask(__name__)

# Global state
wb = None
client_list: List[str] = []

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def read_workbook() -> None:
    """Load workbook and extract sorted client list."""
    global wb, client_list
    try:
        wb = load_workbook(DB_FILE, read_only=False, data_only=True)
        client_list = wb.sheetnames
        client_list.remove(MAIN_SHEET)
        client_list.sort()
        logger.info(f"Workbook loaded. Clients: {len(client_list)}")
    except FileNotFoundError:
        logger.error(f"Database file not found: {DB_FILE}")
        raise
    except Exception as e:
        logger.error(f"Error reading workbook: {e}")
        raise

def is_authorized(chat_id: int) -> bool:
    """Check if user is in whitelist."""
    return chat_id in WHITELIST

def send_to_admin(message: types.Message, content: str, parse_mode: str = 'HTML') -> None:
    """Log action to admin."""
    try:
        if ADMIN_ID:
            bot.send_message(
                ADMIN_ID,
                content,
                parse_mode=parse_mode
            )
    except Exception as e:
        logger.error(f"Failed to notify admin: {e}")

def extract_isins(worksheet) -> List[str]:
    """Extract ISIN codes from worksheet column A."""
    isins = []
    isin_pattern = r"[a-zA-Z]{2}[a-zA-Z0-9]{10}"
    for row in worksheet['A']:
        if row.value and re.search(isin_pattern, str(row.value)):
            isins.append(row.value)
    return isins

# ============================================================================
# BOT MESSAGE HANDLERS
# ============================================================================

@bot.message_handler(commands=['start'])
def handle_start(message: types.Message) -> None:
    """Handle /start command."""
    logger.info(f"Start command from user {message.chat.id}")
    
    if is_authorized(message.chat.id):
        bot.send_message(message.chat.id, "Привет! Справка - /help")
    
    send_to_admin(
        message,
        f"Received /start command from: @{message.chat.username} (id: <code>{message.chat.id}</code>)"
    )

@bot.message_handler(commands=['help'])
def handle_help(message: types.Message) -> None:
    """Handle /help command."""
    logger.info(f"Help command from user {message.chat.id}")
    
    if is_authorized(message.chat.id):
        help_text = (
            "/update - обновить все цены и прислать базу\n"
            "/doc - прислать автобазу (без обновления)\n"
            "/clients - список клиентов в автобазе\n"
            "/report - прислать отчет по клиенту\n\n"
            "Чтобы обновить базу - пришли auto_base.xlsx"
        )
        bot.send_message(message.chat.id, help_text)
    
    send_to_admin(message, f"Received /help command from: @{message.chat.username} (id: <code>{message.chat.id}</code>)")

@bot.message_handler(commands=['doc'])
def handle_doc(message: types.Message) -> None:
    """Send database file without updates."""
    logger.info(f"Doc command from user {message.chat.id}")
    
    if is_authorized(message.chat.id):
        try:
            with open(DB_FILE, 'rb') as f:
                bot.send_document(message.chat.id, f)
        except FileNotFoundError:
            bot.send_message(message.chat.id, "❌ Database file not found")
            logger.error("Database file not found")
    
    send_to_admin(message, f"auto_base.xlsx requested by: @{message.chat.username} (id: <code>{message.chat.id}</code>)")

@bot.message_handler(commands=['update'])
def handle_update(message: types.Message) -> None:
    """Update all prices and send updated database."""
    logger.info(f"Update command from user {message.chat.id}")
    
    if not is_authorized(message.chat.id):
        return
    
    try:
        bot.send_message(message.chat.id, "Обновляю цены...")
        bot.send_chat_action(message.chat.id, "upload_document")
        
        # Load fresh workbook
        wb_update = load_workbook(DB_FILE, read_only=False, data_only=False)
        ws = wb_update[MAIN_SHEET]
        isins = extract_isins(ws)
        
        rows = list(range(1, ws.max_row + 1))
        
        # Update prices concurrently
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(update_price_row, ws, row, isins): row for row in rows}
            for future in futures:
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"Error updating row: {e}")
        
        # Save and send
        output_file = os.path.join(DATA_DIR, 'auto_base_updated.xlsx')
        wb_update.save(output_file)
        
        with open(output_file, 'rb') as f:
            bot.send_document(message.chat.id, f)
        
        os.remove(output_file)
        wb_update.close()
        
    except Exception as e:
        logger.error(f"Error in update command: {e}")
        bot.send_message(message.chat.id, f"❌ Error: {str(e)}")
    
    send_to_admin(message, f"Received /update command from: @{message.chat.username} (id: <code>{message.chat.id}</code>)")

@bot.message_handler(commands=['clients'])
def handle_clients(message: types.Message) -> None:
    """Send list of all clients."""
    logger.info(f"Clients command from user {message.chat.id}")
    
    if is_authorized(message.chat.id):
        bot.send_message(message.chat.id, ", ".join(client_list))
    
    send_to_admin(message, f"Received /clients command from: @{message.chat.username} (id: <code>{message.chat.id}</code>)")

@bot.message_handler(commands=['report'])
def handle_report(message: types.Message) -> None:
    """Request client report with inline keyboard."""
    logger.info(f"Report command from user {message.chat.id}")
    
    if not is_authorized(message.chat.id):
        return
    
    markup = types.ReplyKeyboardMarkup()
    clients_with_cancel = client_list + ["Отмена"]
    
    # Add buttons in pairs
    for i in range(0, len(clients_with_cancel), 2):
        if i + 1 < len(clients_with_cancel):
            markup.add(clients_with_cancel[i], clients_with_cancel[i + 1])
        else:
            markup.add(clients_with_cancel[i])
    
    bot.send_message(message.chat.id, "Чей отчет нужен?", reply_markup=markup)
    
    send_to_admin(message, f"Received /report command from: @{message.chat.username} (id: <code>{message.chat.id}</code>)")

@bot.message_handler(regexp="Отмена")
def handle_cancel(message: types.Message) -> None:
    """Cancel report selection."""
    if is_authorized(message.chat.id):
        markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, "Бывает", reply_markup=markup)

@bot.message_handler(content_types=['document'])
def handle_document(message: types.Message) -> None:
    """Handle file uploads - accept only auto_base.xlsx."""
    file_name = message.document.file_name
    logger.info(f"Document received: {file_name} from user {message.chat.id}")
    
    if not is_authorized(message.chat.id):
        return
    
    if file_name != 'auto_base.xlsx':
        bot.send_message(message.chat.id, "🤨 Проверь имя файла, должно быть: auto_base.xlsx")
        return
    
    try:
        bot.send_chat_action(message.chat.id, "typing")
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        output_path = os.path.join(DATA_DIR, file_name)
        with open(output_path, 'wb') as f:
            f.write(downloaded_file)
        
        read_workbook()
        bot.send_message(message.chat.id, "База обновлена 👍")
        logger.info(f"Database updated successfully")
        
    except Exception as e:
        logger.error(f"Error updating database: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка при обновлении: {str(e)}")
    
    send_to_admin(message, f"Received auto_base.xlsx from: @{message.chat.username} (id: <code>{message.chat.id}</code>)")

@bot.message_handler(func=lambda message: str(message.text).title() in client_list)
def handle_client_selection(message: types.Message) -> None:
    """Generate report for selected client."""
    logger.info(f"Report request for: {message.text} from user {message.chat.id}")
    
    if not is_authorized(message.chat.id):
        return
    
    try:
        client_name = message.text.title()
        bot.send_message(message.chat.id, "Обновляю цены и формирую отчет...")
        bot.send_chat_action(message.chat.id, "upload_document")
        
        # Load workbook
        read_workbook()
        isins = extract_isins(wb[client_name])
        
        ws_main = wb[MAIN_SHEET]
        ws_client = wb[client_name]
        
        rows = list(range(1, ws_main.max_row + 1))
        
        # Update prices concurrently
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {
                executor.submit(update_price_row, ws_main, row, isins, ws_client, client_name): row 
                for row in rows
            }
            for future in futures:
                try:
                    future.result()
                except Exception as e:
                    logger.warning(f"Error updating row: {e}")
        
        # Clean up and save report
        for sheet_name in client_list:
            if sheet_name != client_name:
                del wb[sheet_name]
        del wb[MAIN_SHEET]
        
        output_path = os.path.join(DATA_DIR, f"{client_name}.xlsx")
        wb.save(output_path)
        
        with open(output_path, 'rb') as f:
            bot.send_document(message.chat.id, f)
        
        os.remove(output_path)
        read_workbook()
        
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка при формировании отчета: {str(e)}")
    
    send_to_admin(
        message,
        f"Received report request for {message.text} by: @{message.chat.username} (id: <code>{message.chat.id}</code>)"
    )

@bot.message_handler(func=lambda m: True)
def handle_any_message(message: types.Message) -> None:
    """Log any other message from users."""
    logger.info(f"Message from {message.chat.id}: {message.text[:50]}")
    send_to_admin(
        message,
        f"Received message from: @{message.chat.username} (id: <code>{message.chat.id}</code>):\n{message.text}"
    )

# ============================================================================
# PRICE UPDATE LOGIC
# ============================================================================

def update_price_row(
    ws,
    row: int,
    isins: List[str],
    ws_client=None,
    client_name: str = None
) -> None:
    """Update price for a single row from web sources."""
    try:
        if ws[f'A{row}'].value not in isins:
            return
        
        isin = ws[f'A{row}'].value
        url = ws[f'L{row}'].value
        
        if not url:
            return
        
        bid = scrape_price(url)
        
        if bid == 0:
            logger.warning(f"Unable to update price for ISIN: {isin}")
            ws[f'G{row}'].value = ws[f'G{row}'].value.strftime("%d.%m.%Y") if ws[f'G{row}'].value else None
            ws[f'H{row}'].fill = PatternFill(fgColor='ff5c5c', fill_type='solid')
        elif bid > 40:
            ws[f'G{row}'].value = date.today().strftime("%d.%m.%Y")
            ws[f'H{row}'].value = bid
            ws[f'H{row}'].fill = PatternFill(fgColor='67d948', fill_type='solid')
        else:
            ws[f'H{row}'].fill = PatternFill(fgColor='ffe68a', fill_type='solid')
        
        ws[f'J{row}'].value = float(ws[f'J{row}'].value) if ws[f'J{row}'].value else 0
        ws[f'K{row}'].value = float(ws[f'K{row}'].value) if ws[f'K{row}'].value else 0
        ws[f'C{row}'].value = ws[f'C{row}'].value.strftime("%d.%m.%Y") if ws[f'C{row}'].value else None
        
        # Update client sheet if provided
        if ws_client and client_name:
            for client_row in ws_client['A']:
                if client_row.value == isin:
                    update_client_row(ws, ws_client, row, client_row.row)
                    
    except Exception as e:
        logger.error(f"Error updating price for row {row}: {e}")

def scrape_price(url: str) -> float:
    """Scrape price from URL using appropriate parser."""
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'lxml')
        bid = 0
        
        # boerse-online.de
        if "boerse-online" in url:
            spans = soup.findAll("span", class_=re.compile("kurs aktiendetail"))
            if spans:
                bid = float("".join(re.findall(r'\b\d+\S+\b', spans[0].text.replace(",", "."))))
        
        # boerse-berlin.com
        elif "boerse-berlin" in url or "berliner-boerse" in url:
            spans = soup.findAll("span", re.compile("_bid$|_last"))
            if spans:
                bid = float(spans[-1].text.replace(",", "."))
        
        # oblible
        elif "oblible" in url:
            td = soup.find("td", text="Market price")
            if td:
                next_td = td.find_next_sibling("td", class_=None)
                matches = re.findall(r'\b\d+\S+\b', next_td.text) if next_td else []
                bid = float(matches[0]) if matches else 0
        
        # finanzen.de
        elif "finanzen" in url:
            div = soup.find("div", class_=re.compile("col-xs-5 col-sm-4 text-sm-right text-nowrap$"))
            if div:
                bid = float(div.text[:-1].replace(",", "."))
        
        return bid
        
    except requests.RequestException as e:
        logger.error(f"Request error for {url}: {e}")
        return 0
    except (ValueError, AttributeError) as e:
        logger.error(f"Parse error for {url}: {e}")
        return 0
    except Exception as e:
        logger.error(f"Unexpected error scraping {url}: {e}")
        return 0

def update_client_row(ws: object, ws_client: object, ws_row: int, client_row: int) -> None:
    """Update client sheet row with calculated values."""
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_client[f'{col}{client_row}'].value = ws[f'{col}{ws_row}'].value
    
    # Financial calculations
    ws_client[f'L{client_row}'].value = ws_client[f'J{client_row}'].value * ws_client[f'K{client_row}'].value / 100
    ws_client[f'M{client_row}'].value = ws_client[f'F{client_row}'].value / ws_client[f'K{client_row}'].value
    ws_client[f'N{client_row}'].value = ws_client[f'H{client_row}'].value * ws_client[f'J{client_row}'].value / 100
    ws_client[f'O{client_row}'].value = ws_client[f'N{client_row}'].value - ws_client[f'L{client_row}'].value
    ws_client[f'P{client_row}'].value = ws_client[f'O{client_row}'].value / ws_client[f'L{client_row}'].value if ws_client[f'L{client_row}'].value else 0
    
    days_held = (datetime.today() - ws_client[f'I{client_row}'].value).days if ws_client[f'I{client_row}'].value else 0
    ws_client[f'Q{client_row}'].value = days_held
    ws_client[f'R{client_row}'].value = days_held / 365 * 12 if days_held else 0
    ws_client[f'S{client_row}'].value = ws_client[f'M{client_row}'].value * ws_client[f'L{client_row}'].value / 365 * days_held if days_held else 0
    ws_client[f'T{client_row}'].value = ws_client[f'S{client_row}'].value + ws_client[f'O{client_row}'].value
    ws_client[f'U{client_row}'].value = ws_client[f'T{client_row}'].value / ws_client[f'L{client_row}'].value if ws_client[f'L{client_row}'].value else 0
    ws_client[f'I{client_row}'].value = ws_client[f'I{client_row}'].value.strftime("%d.%m.%Y") if ws_client[f'I{client_row}'].value else None

# ============================================================================
# SERVER & STARTUP
# ============================================================================

if __name__ == "__main__":
    # Initialize
    read_workbook()
    
    if "HEROKU" in os.environ:
        # Webhook mode for production
        @server.route('/' + TOKEN, methods=['POST'])
        def handle_webhook():
            json_string = request.stream.read().decode('utf-8')
            update = telebot.types.Update.de_json(json_string)
            bot.process_new_updates([update])
            return "OK", 200
        
        @server.route("/")
        def set_webhook():
            bot.remove_webhook()
            webhook_url = os.getenv('WEBHOOK_URL', 'https://securities-helper.herokuapp.com')
            bot.set_webhook(url=f'{webhook_url}/{TOKEN}')
            return "Webhook set", 200
        
        logger.info("Starting in webhook mode")
        server.run(host="0.0.0.0", port=int(os.environ.get('PORT', 5000)))
    else:
        # Polling mode for local development
        logger.info("Starting in polling mode")
        bot.remove_webhook()
        bot.polling(none_stop=True, skip_pending=True)
